from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter, OrderingFilter
from .models import Process
from .serializers import (
    ProcessSerializer,
    ProcessListSerializer,
    ProcessCreateUpdateSerializer
)
from parties.models import Party
from parties.serializers import PartySerializer
import openpyxl
from django.http import HttpResponse
from django.utils import timezone


class ProcessViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing legal processes.
    """
    queryset = Process.objects.all()
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['process_class', 'judge']
    search_fields = ['process_number', 'subject', 'judge']
    ordering_fields = ['process_number', 'created_at', 'updated_at']
    ordering = ['-created_at']

    def get_serializer_class(self):
        """Return appropriate serializer class based on action."""
        if self.action in ['create', 'update', 'partial_update']:
            return ProcessCreateUpdateSerializer
        elif self.action == 'list':
            return ProcessListSerializer
        return ProcessSerializer

    @action(detail=True, methods=['get'])
    def parties(self, request, pk=None):
        """Get all parties for a specific process."""
        process = self.get_object()
        parties = Party.objects.filter(process=process)
        serializer = PartySerializer(parties, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """Export processes data to Excel file."""
        processes = self.filter_queryset(self.get_queryset())
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Processes"
        
        # Add headers
        headers = [
            'Process Number', 'Class', 'Subject', 'Judge', 
            'Parties Count', 'Created At'
        ]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Add data
        for row, process in enumerate(processes, 2):
            ws.cell(row=row, column=1, value=process.process_number)
            ws.cell(row=row, column=2, value=process.process_class)
            ws.cell(row=row, column=3, value=process.subject)
            ws.cell(row=row, column=4, value=process.judge)
            ws.cell(row=row, column=5, value=process.parties_count)
            ws.cell(row=row, column=6, value=process.created_at.strftime('%Y-%m-%d %H:%M'))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=processes_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        wb.save(response)
        return response

    @action(detail=False, methods=['post'])
    def bulk_create(self, request):
        """Create multiple processes at once."""
        serializer = ProcessCreateUpdateSerializer(data=request.data, many=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
